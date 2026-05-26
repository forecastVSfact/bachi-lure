"""
Convert lure Excel (v3 format) to import-ready CSV.
Usage: python scripts/xlsx-to-csv.py [input.xlsx] [output.csv]
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl

CANONICAL_SPEED = frozenset({"dead_slow", "slow", "medium", "all"})
CANONICAL_BACHI = frozenset({"river", "harbor", "kurukuru", "bottom"})
BACHI_ALIASES = {
    "harbor_drift": "harbor",
    "harbor_wave": "harbor",
}


def parse_header(cell) -> str | None:
    if cell is None:
        return None
    text = str(cell).replace("\n", " ")
    match = re.search(r"\(([^)]+)\)\s*$", text)
    return match.group(1) if match else None


def normalize_speed_range(raw) -> str:
    if raw is None or str(raw).strip() == "":
        return "all"
    text = str(raw).replace("、", ",").replace("　", "").strip()
    parts = [p.strip() for p in text.split(",") if p.strip()]
    canonical: set[str] = set()
    for part in parts:
        key = part.lower().replace("-", "_")
        if key == "deadslow":
            key = "dead_slow"
        elif key in ("fast", "medium_fast"):
            key = "medium"
        if key in CANONICAL_SPEED:
            canonical.add(key)
    if not canonical:
        return "all"
    if len(canonical) == 1:
        return next(iter(canonical))
    return "all"


def _format_rating(raw) -> str:
    if raw is None:
        return ""
    text = str(raw).strip()
    if text in ("", "-", "―", "ー"):
        return ""
    try:
        value = int(float(text))
    except ValueError:
        return ""
    if 1 <= value <= 5:
        return str(value)
    return ""


def _format_image_urls(raw) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    text = str(raw).replace("、", ",").replace("\n", ",")
    return ",".join(part.strip() for part in text.split(",") if part.strip())


def normalize_bachi_types(raw) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    text = str(raw).replace("、", ",")
    values: list[str] = []
    for part in text.split(","):
        key = part.strip().lower()
        key = BACHI_ALIASES.get(key, key)
        if key in CANONICAL_BACHI and key not in values:
            values.append(key)
    return ",".join(values)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    default_in = root / "luredatabase" / "05131403_lures-data-v3.xlsx"
    default_out = root / "luredatabase" / "lures.csv"

    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_in
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else default_out

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    fields: list[str] = []
    for cell in header_row:
        field = parse_header(cell)
        fields.append(field if field else "comment")

    columns = [
        "name",
        "maker",
        "size_mm",
        "weight_g",
        "price_yen",
        "hook_size",
        "lure_type",
        "range_min_cm",
        "range_max_cm",
        "swim_posture",
        "action",
        "speed_range",
        "casting_distance",
        "bachi_types",
        "youtube_url",
        "amazon_url",
        "rakuten_url",
        "rating",
        "comment",
    ]
    if "image_urls" in fields:
        columns.insert(columns.index("comment"), "image_urls")
    elif "image_url" in fields:
        columns.insert(columns.index("comment"), "image_urls")
    index = {name: fields.index(name) for name in fields if name in fields}

    rows_out: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[index["name"]]:
            continue

        def cell(name: str):
            return row[index[name]] if name in index else None

        rows_out.append(
            {
                "name": str(cell("name")).strip(),
                "maker": str(cell("maker")).strip(),
                "size_mm": "" if cell("size_mm") is None else str(cell("size_mm")),
                "weight_g": "" if cell("weight_g") is None else str(cell("weight_g")),
                "price_yen": "" if cell("price_yen") is None else str(int(cell("price_yen"))),
                "hook_size": "" if cell("hook_size") is None else str(cell("hook_size")),
                "lure_type": str(cell("lure_type")).strip(),
                "range_min_cm": "" if cell("range_min_cm") is None else str(cell("range_min_cm")),
                "range_max_cm": "" if cell("range_max_cm") is None else str(cell("range_max_cm")),
                "swim_posture": str(cell("swim_posture")).strip(),
                "action": "" if cell("action") is None else str(cell("action")).strip(),
                "speed_range": normalize_speed_range(cell("speed_range")),
                "casting_distance": str(cell("casting_distance")).strip(),
                "bachi_types": normalize_bachi_types(cell("bachi_types")),
                "youtube_url": "" if not cell("youtube_url") else str(cell("youtube_url")).strip(),
                "amazon_url": "" if not cell("amazon_url") else str(cell("amazon_url")).strip(),
                "rakuten_url": "" if not cell("rakuten_url") else str(cell("rakuten_url")).strip(),
                "rating": _format_rating(cell("rating")),
                "image_urls": _format_image_urls(cell("image_urls") if "image_urls" in index else cell("image_url") if "image_url" in index else None),
                "comment": "" if not cell("comment") else str(cell("comment")).strip(),
            }
        )

    wb.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out)} rows -> {output_path}")


if __name__ == "__main__":
    main()
